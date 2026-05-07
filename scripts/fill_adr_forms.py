#!/usr/bin/env python3
"""
fill_adr_forms.py
═════════════════
Generate filled ADR Reporting Form 1.4 PDFs from ICSR Excel datasets.

Supports two datasets out of the box:
  • data/ADRA_Synthetic_Evaluation_Dataset.xlsx  →  sheet ADRA_ICSR_Synthetic
  • data/CDSCO_AI_Datasets.xlsx                  →  sheet 1_SAE_ICSRs

Requirements:
    pip install pypdf openpyxl

Usage:
    # Fill 50 forms from each dataset (default)
    python scripts/fill_adr_forms.py

    # Fill 100 forms from one specific file
    python scripts/fill_adr_forms.py --input data/CDSCO_AI_Datasets.xlsx --sheet 1_SAE_ICSRs --limit 100

    # Discover all field names in the template
    python scripts/fill_adr_forms.py --discover

Output:
    output/filled_forms/<source_tag>/<ICSR_ID>.pdf
"""

import argparse
import logging
import sys
from pathlib import Path

# Suppress benign pypdf "Ignoring wrong pointing object" warnings from template PDF
logging.getLogger("pypdf").setLevel(logging.ERROR)

try:
    from pypdf import PdfReader, PdfWriter
    from pypdf.generic import NameObject
except ImportError:
    print("pypdf not found. Install with: pip install pypdf")
    sys.exit(1)

try:
    import openpyxl
except ImportError:
    print("openpyxl not found. Install with: pip install openpyxl")
    sys.exit(1)

# ─── Paths ────────────────────────────────────────────────────────────────────

SCRIPT_DIR = Path(__file__).parent
ROOT = SCRIPT_DIR.parent
DATA = ROOT / "data"
PDF_TEMPLATE = DATA / "ADR_Reporting_Form_1.4_Version.pdf"
DEFAULT_OUTPUT = ROOT / "output" / "filled_forms"

DEFAULT_DATASETS = [
    (DATA / "ADRA_Synthetic_Evaluation_Dataset.xlsx", "ADRA_ICSR_Synthetic", "ADRA"),
    (DATA / "CDSCO_AI_Datasets.xlsx",                 "1_SAE_ICSRs",         "CDSCO"),
]

# ─── Field mapping ────────────────────────────────────────────────────────────
# Field names are generic in this PDF ("Text Field0" etc.).
# Mapping derived from reading-order analysis (page, y descending, x ascending).
# Cross-referenced with the standard PvPI ADR Reporting Form 1.4 layout.

TEXT_FIELDS = {
    # Report date
    "Text Field3":   "report_date",
    # Patient information
    "Text Field0":   "patient_initials",
    "Text Field2":   "country",
    "Text Field4":   "region",
    "Text Field126": "patient_age",
    "Text Field5":   "patient_weight",
    "Text Field6":   "patient_dob",
    # Adverse reaction description (large text box)
    "Text Field7":   "adverse_reaction",
    # Other relevant medical history / concomitant
    "Text Field9":   "other_info",
    # Onset date (single field — full date)
    "Text Field48":  "onset_date",
    # Narrative / additional information (large text box)
    "Text Field8":   "narrative",
    # ── Suspect drug table (drug 1) ──
    "Text Field10":  "drug1_name",
    "Text Field11":  "drug1_manufacturer",
    "Text Field12":  "drug1_batch",
    "Text Field13":  "drug1_expiry",
    "Text Field14":  "drug1_dose",
    "Text Field15":  "drug1_dose_unit",
    "Text Field16":  "drug1_route",
    "Text Field17":  "drug1_freq",
    "Text Field18":  "drug1_start",
    "Text Field19":  "drug1_stop",
    "Text Field20":  "drug1_indication",
    # ── Suspect drug table (drug 2 / concomitant 1) ──
    "Text Field49":  "drug2_name",
    "Text Field52":  "drug2_manufacturer",
    "Text Field55":  "drug2_batch",
    "Text Field58":  "drug2_expiry",
    "Text Field61":  "drug2_dose",
    "Text Field64":  "drug2_route",
    "Text Field67":  "drug2_freq",
    "Text Field70":  "drug2_start",
    "Text Field85":  "drug2_stop",
    "Text Field88":  "drug2_indication",
    # ── Suspect drug table (drug 3 / concomitant 2) ──
    "Text Field50":  "drug3_name",
    "Text Field53":  "drug3_manufacturer",
    "Text Field56":  "drug3_batch",
    "Text Field59":  "drug3_expiry",
    "Text Field62":  "drug3_dose",
    "Text Field65":  "drug3_route",
    "Text Field68":  "drug3_freq",
    "Text Field71":  "drug3_start",
    "Text Field86":  "drug3_stop",
    "Text Field89":  "drug3_indication",
    # ── Reaction onset/recovery date parts (separate columns) ──
    "Text Field31":  "reaction_onset_day",
    "Text Field32":  "reaction_onset_month",
    "Text Field33":  "reaction_onset_year",
    "Text Field34":  "reaction_onset_time",
    "Text Field35":  "reaction_recovery_date",
    "Text Field36":  "reaction_outcome",
    "Text Field37":  "causality",
    # ── Concomitant drug section (separate table) ──
    "Text Field112": "con1_name",
    "Text Field114": "con1_dose",
    "Text Field116": "con1_route",
    "Text Field118": "con1_freq",
    "Text Field120": "con1_start",
    "Text Field122": "con1_stop",
    "Text Field124": "con1_indication",
    "Text Field113": "con2_name",
    "Text Field115": "con2_dose",
    "Text Field117": "con2_route",
    "Text Field119": "con2_freq",
    "Text Field121": "con2_start",
    "Text Field123": "con2_stop",
    "Text Field125": "con2_indication",
    # ── Reporter information ──
    "Text Field38":  "reporter_notes",
    "Text Field39":  "reporter_name_address",
    "Text Field40":  "reporter_full_address",
    "Text Field41":  "reporter_city",
    "Text Field42":  "reporter_pin",
    "Text Field43":  "reporter_state",
    "Text Field44":  "reporter_phone",
    "Text Field45":  "reporter_email",
    "Text Field46":  "reporter_date",
    "Text Field47":  "reporter_qualification",
}

CHECKBOX_FIELDS = {
    # Reporter type
    "Check Box16": "cb_consumer",
    "Check Box1":  "cb_hcp",
    # Patient sex
    "Check Box0":  "cb_male",
    "Check Box17": "cb_female",
    "Check Box18": "cb_sex_unknown",
    # Seriousness criteria
    "Check Box2":  "cb_fatal",
    "Check Box3":  "cb_life_threatening",
    "Check Box4":  "cb_hospitalisation",
    "Check Box5":  "cb_disability",
    "Check Box6":  "cb_congenital",
    "Check Box7":  "cb_other_serious",
    # Outcome
    "Check Box8":  "cb_recovered",
    "Check Box9":  "cb_recovering",
    "Check Box10": "cb_not_recovered",
    "Check Box11": "cb_outcome_unknown",
    # Dechallenge
    "Check Box12": "cb_dechallenge_yes",
    "Check Box13": "cb_dechallenge_no",
    "Check Box14": "cb_dechallenge_unknown",
    # Rechallenge
    "Check Box15": "cb_rechallenge_positive",
    # Check Box16 also used for rechallenge negative in some form versions:
    # "Check Box16": "cb_rechallenge_negative",
}


# ─── Row → field value builder ────────────────────────────────────────────────

def _str(val) -> str:
    """Safe string conversion from Excel cell value."""
    if val is None:
        return ""
    return str(val).strip()


def _split_date(date_str: str):
    """Return (year, month, day) from YYYY-MM-DD or DD/MM/YYYY."""
    s = _str(date_str)
    if "-" in s:
        parts = s.split("-")
        if len(parts) == 3 and len(parts[0]) == 4:
            return parts[0], parts[1], parts[2]   # YYYY-MM-DD
    if "/" in s:
        parts = s.split("/")
        if len(parts) == 3:
            return parts[2], parts[1], parts[0]   # DD/MM/YYYY → year, month, day
    return "", "", s


def build_field_values(row: dict) -> tuple[dict, dict]:
    """
    Build text-field and checkbox dicts from an ICSR row.
    Returns (text_values, checkbox_values) both keyed by semantic name.
    """
    # Core fields
    reporter_type   = _str(row.get("Reporter_Type", ""))
    sex             = _str(row.get("Patient_Sex", ""))
    age             = _str(row.get("Patient_Age", ""))
    weight          = _str(row.get("Patient_Weight_kg", ""))
    suspect_drug    = _str(row.get("Suspect_Drug", ""))
    dose            = _str(row.get("Drug_Dose_mg", ""))
    route           = _str(row.get("Drug_Route", ""))
    freq            = _str(row.get("Dose_Frequency", ""))
    indication      = _str(row.get("Indication", ""))
    meddra_pt       = _str(row.get("MedDRA_PT", ""))
    meddra_soc      = _str(row.get("MedDRA_SOC", ""))
    meddra_llt      = _str(row.get("MedDRA_LLT", ""))
    seriousness     = _str(row.get("SAE_Seriousness_Criteria", ""))
    outcome         = _str(row.get("Outcome", ""))
    causality       = _str(row.get("Causality_Assessment", ""))
    dechallenge     = _str(row.get("Dechallenge", "Unknown"))
    rechallenge     = _str(row.get("Rechallenge", "Unknown"))
    concomitant     = _str(row.get("Concomitant_Drugs", ""))
    narrative       = _str(row.get("Narrative", ""))
    region          = _str(row.get("Region", row.get("Site_ID", "")))
    report_date     = _str(row.get("Report_Date", ""))
    onset_date      = _str(row.get("Onset_Date", ""))

    onset_year, onset_month, onset_day = _split_date(onset_date)

    # Concomitant drug list (can be comma-separated or a count integer)
    con_drugs = []
    if concomitant and not concomitant.isdigit():
        con_drugs = [d.strip() for d in concomitant.replace(";", ",").split(",") if d.strip()]

    # Adverse reaction description (combine MedDRA terms)
    reaction_desc = meddra_pt
    if meddra_llt and meddra_llt != meddra_pt:
        reaction_desc += f" ({meddra_llt})"
    if meddra_soc:
        reaction_desc += f"\nMedDRA SOC: {meddra_soc}"

    text = {
        "report_date":           report_date,
        "patient_initials":      "P.P.",        # always anonymised
        "country":               "India",
        "region":                region,
        "patient_age":           age,
        "patient_weight":        weight,
        "patient_dob":           "",
        "adverse_reaction":      reaction_desc,
        "other_info":            f"Indication: {indication}",
        "onset_date":            onset_date,
        "narrative":             narrative[:1000],   # truncate for field size
        # Drug 1 — suspect drug
        "drug1_name":            suspect_drug,
        "drug1_manufacturer":    "",
        "drug1_batch":           "",
        "drug1_expiry":          "",
        "drug1_dose":            dose,
        "drug1_dose_unit":       "mg",
        "drug1_route":           route,
        "drug1_freq":            freq,
        "drug1_start":           "",
        "drug1_stop":            "",
        "drug1_indication":      indication,
        # Drug 2 — first concomitant
        "drug2_name":            con_drugs[0] if len(con_drugs) > 0 else "",
        "drug2_manufacturer":    "",
        "drug2_batch":           "",
        "drug2_expiry":          "",
        "drug2_dose":            "",
        "drug2_route":           "",
        "drug2_freq":            "",
        "drug2_start":           "",
        "drug2_stop":            "",
        "drug2_indication":      "",
        # Drug 3 — second concomitant
        "drug3_name":            con_drugs[1] if len(con_drugs) > 1 else "",
        "drug3_manufacturer":    "",
        "drug3_batch":           "",
        "drug3_expiry":          "",
        "drug3_dose":            "",
        "drug3_route":           "",
        "drug3_freq":            "",
        "drug3_start":           "",
        "drug3_stop":            "",
        "drug3_indication":      "",
        # Reaction onset/recovery
        "reaction_onset_day":    onset_day,
        "reaction_onset_month":  onset_month,
        "reaction_onset_year":   onset_year,
        "reaction_onset_time":   "",
        "reaction_recovery_date":"",
        "reaction_outcome":      outcome,
        "causality":             causality,
        # Concomitant drug section (additional table rows)
        "con1_name":             con_drugs[0] if len(con_drugs) > 0 else "",
        "con1_dose":             "",
        "con1_route":            "",
        "con1_freq":             "",
        "con1_start":            "",
        "con1_stop":             "",
        "con1_indication":       "",
        "con2_name":             con_drugs[1] if len(con_drugs) > 1 else "",
        "con2_dose":             "",
        "con2_route":            "",
        "con2_freq":             "",
        "con2_start":            "",
        "con2_stop":             "",
        "con2_indication":       "",
        # Reporter
        "reporter_notes":        narrative[:300],
        "reporter_name_address": "",
        "reporter_full_address": region,
        "reporter_city":         region,
        "reporter_pin":          "",
        "reporter_state":        "",
        "reporter_phone":        "",
        "reporter_email":        "",
        "reporter_date":         report_date,
        "reporter_qualification":reporter_type,
    }

    def yn(val: str, *keywords) -> bool:
        v = val.lower()
        return any(kw in v for kw in keywords)

    checks = {
        "cb_consumer":           yn(reporter_type, "consumer", "patient"),
        "cb_hcp":                not yn(reporter_type, "consumer", "patient"),
        "cb_male":               sex.lower() == "male",
        "cb_female":             sex.lower() == "female",
        "cb_sex_unknown":        sex.lower() not in ("male", "female"),
        # Seriousness — maps all common variants
        "cb_fatal":              yn(seriousness, "death", "fatal"),
        "cb_life_threatening":   yn(seriousness, "life-threatening", "life threatening"),
        "cb_hospitalisation":    yn(seriousness, "hospitalisation", "hospitali"),
        "cb_disability":         yn(seriousness, "disability", "incapacity"),
        "cb_congenital":         yn(seriousness, "congenital"),
        "cb_other_serious":      yn(seriousness, "other medically"),
        # Outcome
        "cb_recovered":          yn(outcome, "recovered/resolved", "recovered"),
        "cb_recovering":         yn(outcome, "recovering", "resolving"),
        "cb_not_recovered":      yn(outcome, "not recovered", "not resolved", "ongoing"),
        "cb_outcome_unknown":    yn(outcome, "unknown", "not available"),
        # Dechallenge
        "cb_dechallenge_yes":    yn(dechallenge, "yes", "abated", "improved", "positive"),
        "cb_dechallenge_no":     yn(dechallenge, "no", "did not", "not done"),
        "cb_dechallenge_unknown":yn(dechallenge, "unknown", "n/a", "not applicable"),
        # Rechallenge
        "cb_rechallenge_positive":yn(rechallenge, "yes", "positive", "recurred"),
    }

    return text, checks


# ─── PDF filler ───────────────────────────────────────────────────────────────

def _get_checkbox_on_value(annot_obj) -> str:
    """Return the /On appearance state name for a checkbox annotation."""
    ap = annot_obj.get("/AP")
    if ap and hasattr(ap, "get"):
        n = ap.get("/N")
        if n and hasattr(n, "keys"):
            on_vals = [k for k in n.keys() if k not in ("/Off", "Off")]
            if on_vals:
                return on_vals[0]
    return "/Yes"


def fill_pdf(template_path: Path, text_vals: dict, check_vals: dict, out_path: Path) -> None:
    """Fill the PDF form and write to out_path."""
    reader = PdfReader(str(template_path))
    writer = PdfWriter()
    writer.append(reader)

    # Build text field update dict: PDF field name → value
    text_updates = {}
    for pdf_name, semantic in TEXT_FIELDS.items():
        val = text_vals.get(semantic, "")
        if val:
            text_updates[pdf_name] = str(val)

    # Apply text fields
    if text_updates and writer.pages:
        writer.update_page_form_field_values(writer.pages[0], text_updates)

    # Apply checkboxes by walking annotations
    if writer.pages:
        page = writer.pages[0]
        annots = page.get("/Annots")
        if annots:
            for annot_ref in annots:
                try:
                    annot = annot_ref.get_object()
                except Exception:
                    continue
                subtype = annot.get("/Subtype")
                field_type = annot.get("/FT")
                field_name = annot.get("/T")

                if subtype != "/Widget" or field_type != "/Btn" or not field_name:
                    continue

                semantic = CHECKBOX_FIELDS.get(field_name)
                if semantic is None:
                    continue

                should_check = check_vals.get(semantic, False)
                if should_check:
                    on_val = _get_checkbox_on_value(annot)
                    annot[NameObject("/V")] = NameObject(on_val)
                    annot[NameObject("/AS")] = NameObject(on_val)
                else:
                    annot[NameObject("/V")] = NameObject("/Off")
                    annot[NameObject("/AS")] = NameObject("/Off")

    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "wb") as fh:
        writer.write(fh)


# ─── Excel reader ─────────────────────────────────────────────────────────────

def load_rows(excel_path: Path, sheet_name: str, limit: int) -> list[dict]:
    """Load rows from an Excel sheet as list of dicts."""
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        print(f"  Sheet '{sheet_name}' not found in {excel_path.name}. Available: {wb.sheetnames}")
        wb.close()
        return []
    ws = wb[sheet_name]
    rows_iter = ws.iter_rows(values_only=True)
    headers_row = next(rows_iter, None)
    if headers_row is None:
        wb.close()
        return []
    headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(headers_row)]
    result = []
    for raw in rows_iter:
        if len(result) >= limit:
            break
        result.append({headers[i]: (raw[i] if i < len(raw) else None) for i in range(len(headers))})
    wb.close()
    return result


# ─── CLI ─────────────────────────────────────────────────────────────────────

def discover_fields(template: Path) -> None:
    """Print all AcroForm field names, types and positions."""
    reader = PdfReader(str(template))
    fields = reader.get_fields()
    if not fields:
        print("No AcroForm fields found in the PDF.")
        return
    print(f"{'#':>4}  {'Field name':35}  {'Type':12}  Current value")
    for i, (name, field) in enumerate(sorted(fields.items()), 1):
        ftype = _str(field.get("/FT", "?"))
        fval = _str(field.get("/V", ""))
        print(f"{i:>4}  {name:35}  {ftype:12}  {fval}")
    print(f"\nTotal: {len(fields)} fields")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Generate filled ADR Reporting Form 1.4 PDFs from ICSR Excel datasets."
    )
    parser.add_argument("--input",      metavar="FILE",  help="Path to Excel file (overrides defaults)")
    parser.add_argument("--sheet",      metavar="SHEET", help="Sheet name to use with --input")
    parser.add_argument("--limit",      type=int, default=50, metavar="N",
                        help="Max rows to process per dataset (default: 50)")
    parser.add_argument("--output-dir", metavar="DIR", default=str(DEFAULT_OUTPUT),
                        help=f"Output directory (default: {DEFAULT_OUTPUT})")
    parser.add_argument("--discover",   action="store_true",
                        help="Print all PDF field names and exit (no filling)")
    parser.add_argument("--verbose",    action="store_true", help="Print each generated file path")
    args = parser.parse_args()

    if not PDF_TEMPLATE.exists():
        print(f"ERROR: Template PDF not found at {PDF_TEMPLATE}")
        sys.exit(1)

    if args.discover:
        discover_fields(PDF_TEMPLATE)
        return

    output_root = Path(args.output_dir)
    total_ok = 0
    total_err = 0

    # Build dataset list
    if args.input:
        sheet = args.sheet or "Sheet1"
        datasets = [(Path(args.input), sheet, Path(args.input).stem)]
    else:
        datasets = [(p, s, tag) for p, s, tag in DEFAULT_DATASETS if p.exists()]

    if not datasets:
        print("No datasets found. Run from the project root or use --input.")
        sys.exit(1)

    for excel_path, sheet_name, tag in datasets:
        print(f"\n{'─'*60}")
        print(f"Dataset : {excel_path.name}")
        print(f"Sheet   : {sheet_name}")
        print(f"Tag     : {tag}")
        rows = load_rows(excel_path, sheet_name, args.limit)
        if not rows:
            print("  No rows loaded — skipping.")
            continue
        print(f"Loaded  : {len(rows)} rows (limit {args.limit})")
        print(f"Output  : {output_root / tag}")

        for i, row in enumerate(rows):
            icsr_id = _str(row.get("ICSR_ID", f"row_{i+1}")).replace("/", "_").replace("\\", "_")
            out_path = output_root / tag / f"{icsr_id}.pdf"
            try:
                text_vals, check_vals = build_field_values(row)
                fill_pdf(PDF_TEMPLATE, text_vals, check_vals, out_path)
                total_ok += 1
                if args.verbose:
                    print(f"  ✓ {out_path.name}")
                elif (i + 1) % 25 == 0 or i == len(rows) - 1:
                    print(f"  Progress: {i+1}/{len(rows)} PDFs generated")
            except Exception as exc:
                total_err += 1
                print(f"  ERROR row {i+1} ({icsr_id}): {exc}")

    print(f"\n{'═'*60}")
    print(f"Generated : {total_ok} PDF(s)")
    print(f"Errors    : {total_err}")
    print(f"Output dir: {output_root}")
    if total_ok:
        print("\nNext steps:")
        print("  1. Upload the generated PDFs via the ADRA Intake page")
        print("  2. Run: npm run evaluate  (to score classifiers against this data)")
        print("  3. Check the Records page for flags, severity, and score distributions")


if __name__ == "__main__":
    main()
